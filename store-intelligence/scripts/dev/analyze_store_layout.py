"""Extract and analyze Brigade Road store layout Excel file."""
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

SI_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = SI_ROOT.parent
XLSX = REPO_ROOT / "Brigade Road - Store layoutc5f5d56.xlsx"
OUT = SI_ROOT / "data" / "layout"
OUT.mkdir(parents=True, exist_ok=True)


def inspect_workbook():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== {name} ===")
        print("Dimensions:", ws.dimensions)
        print("Max row/col:", ws.max_row, ws.max_column)
        print("Merged:", [str(m) for m in ws.merged_cells.ranges])
        images = getattr(ws, "_images", [])
        print("Embedded images:", len(images))
        cells = [
            (c.coordinate, c.value)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
            for c in row
            if c.value is not None and str(c.value).strip()
        ]
        print("Non-empty cells:", cells)


def extract_media():
    with zipfile.ZipFile(XLSX) as z:
        for name in z.namelist():
            if name.startswith("xl/media/"):
                data = z.read(name)
                out = OUT / Path(name).name
                out.write_bytes(data)
                print(f"Extracted {out} ({len(data)} bytes)")


def parse_drawing():
    with zipfile.ZipFile(XLSX) as z:
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8")
    out = OUT / "drawing1.xml"
    out.write_text(drawing, encoding="utf-8")
    print(f"Saved {out} ({len(drawing)} chars)")

    root = ET.fromstring(drawing)
    texts = []
    for t in root.iter("{http://schemas.openxmlformats.org/drawingml/2006/main}t"):
        if t.text and t.text.strip():
            texts.append(t.text.strip())
    print("\nText labels in drawing:", texts[:50])
    print("Total text elements:", len(texts))


if __name__ == "__main__":
    inspect_workbook()
    extract_media()
    parse_drawing()
